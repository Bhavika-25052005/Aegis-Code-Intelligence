import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Backlog"

headers = ["Type", "ID", "Title", "Description", "Parent ID", "Acceptance Criteria"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

data = [
    ["Feature", "104", "Medical Prescriptions",
     "Enable doctors to create, manage, and track patient prescriptions",
     "", ""],

    ["User Story", "207", "Create Prescription",
     "As a doctor, I want to create a prescription for a patient so they can get their medication",
     "104",
     "- Select patient and add one or more medications\n- Specify dosage, frequency, and duration for each medication\n- Prescription saved with doctor's name and timestamp\n- PDF generation of the prescription"],

    ["Task", "319", "Create Prescription database model",
     "Define SQLAlchemy model for Prescription with fields: id, patient_id, doctor_name, created_at, notes. Also create PrescriptionItem model with: id, prescription_id, medication_name, dosage, frequency, duration",
     "207", ""],

    ["Task", "320", "Build POST /api/prescriptions endpoint",
     "Implement prescription creation endpoint that accepts patient_id, doctor_name, notes, and a list of medication items with dosage/frequency/duration. Validate that patient exists.",
     "207", ""],

    ["Task", "321", "Create prescription form Vue component",
     "Build a form component that allows selecting a patient, adding multiple medications with dosage fields, and submitting the prescription. Include add/remove medication row buttons.",
     "207", ""],

    ["User Story", "208", "View Patient Prescriptions",
     "As a doctor, I want to view all prescriptions for a patient so I can review their medication history",
     "104",
     "- List all prescriptions for a patient sorted by date\n- Each prescription shows medications, dosages, and prescribing doctor\n- Expandable detail view for each prescription"],

    ["Task", "322", "Create GET /api/patients/{id}/prescriptions endpoint",
     "Return all prescriptions for a given patient with their items, ordered by creation date descending. Include pagination.",
     "208", ""],

    ["Task", "323", "Build prescription history list component",
     "Create a Vue component showing a timeline of prescriptions with expandable cards showing medication details for each prescription.",
     "208", ""],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    row_type = row_data[0]
    if row_type == "Feature":
        fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    elif row_type == "User Story":
        fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    else:
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_idx in range(1, 7):
        ws.cell(row=row_idx, column=col_idx).fill = fill

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 70
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 60

wb.save("feature2_prescriptions.xlsx")
print("Created docs/feature2_prescriptions.xlsx")
print("  - 1 Feature: Medical Prescriptions")
print("  - 2 User Stories")
print("  - 5 Tasks")
