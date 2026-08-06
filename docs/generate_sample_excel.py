import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Backlog"

headers = ["Type", "ID", "Title", "Description", "Parent ID", "Acceptance Criteria"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

data = [
    # Feature 1
    ["Feature", "101", "Patient Management System",
     "Build a comprehensive patient management module for healthcare providers", "", ""],

    ["User Story", "201", "Patient Registration",
     "As a receptionist, I want to register new patients so their records are stored in the system", "101",
     "- Patient can be registered with name, DOB, contact info, insurance details\n- Duplicate detection based on name + DOB\n- Confirmation message shown after registration"],

    ["Task", "301", "Create Patient model and database schema",
     "Define SQLAlchemy model for Patient with fields: id, first_name, last_name, dob, email, phone, insurance_id, created_at", "201", ""],

    ["Task", "302", "Build POST /api/patients registration endpoint",
     "Implement patient registration endpoint with input validation, duplicate check, and proper error responses", "201", ""],

    ["Task", "303", "Create patient registration form component",
     "Build a Vue form with fields for patient info, client-side validation, and success/error feedback", "201", ""],

    ["Task", "304", "Add duplicate patient detection logic",
     "Implement fuzzy matching on name + DOB to detect potential duplicates before saving", "201", ""],

    ["User Story", "202", "Patient Search",
     "As a doctor, I want to search for patients so I can quickly find their records", "101",
     "- Search by name, DOB, or patient ID\n- Results shown in a paginated table\n- Search is case-insensitive and supports partial matches"],

    ["Task", "305", "Create GET /api/patients/search endpoint",
     "Implement search endpoint with query params for name, dob, id with pagination support", "202", ""],

    ["Task", "306", "Build patient search UI with filters",
     "Create a search bar component with autocomplete and filter options (name, DOB, ID)", "202", ""],

    ["Task", "307", "Add patient results table with pagination",
     "Build a data table showing search results with sortable columns and page navigation", "202", ""],

    ["User Story", "203", "Patient Profile View",
     "As a healthcare provider, I want to view a patient profile so I can see their complete information", "101",
     "- Profile shows personal info, contact details, insurance\n- History of visits displayed in timeline\n- Allergies and medications listed prominently"],

    ["Task", "308", "Create GET /api/patients/{id} detail endpoint",
     "Return full patient record with related visits, allergies, and medications", "203", ""],

    ["Task", "309", "Build patient profile page UI",
     "Create a detailed profile page with tabs for info, history, allergies, and medications", "203", ""],

    # Feature 2
    ["Feature", "102", "Appointment Scheduling",
     "Enable scheduling and management of patient appointments", "", ""],

    ["User Story", "204", "Book Appointment",
     "As a receptionist, I want to book appointments so patients can schedule visits", "102",
     "- Select doctor, date, time slot, and patient\n- Prevent double-booking same doctor at same time\n- Send confirmation notification to patient"],

    ["Task", "310", "Create Appointment model and schema",
     "Define model with fields: id, patient_id, doctor_id, datetime, duration, status, notes", "204", ""],

    ["Task", "311", "Build POST /api/appointments booking endpoint",
     "Implement appointment creation with conflict detection and validation", "204", ""],

    ["Task", "312", "Create appointment booking form UI",
     "Build a form with doctor selector, date picker, time slot picker, and patient search", "204", ""],

    ["Task", "313", "Add appointment conflict detection",
     "Check for overlapping appointments for the same doctor before confirming", "204", ""],

    ["User Story", "205", "View Calendar",
     "As a doctor, I want to view my appointment calendar so I can see my schedule", "102",
     "- Calendar shows daily/weekly/monthly views\n- Appointments color-coded by status\n- Click appointment to see details"],

    ["Task", "314", "Create GET /api/appointments/calendar endpoint",
     "Return appointments for a doctor within a date range, grouped by day", "205", ""],

    ["Task", "315", "Build calendar view component",
     "Create an interactive calendar with day/week/month toggle and appointment cards", "205", ""],

    # Feature 3
    ["Feature", "103", "Notifications System",
     "Implement notification system for appointment reminders and updates", "", ""],

    ["User Story", "206", "Appointment Reminders",
     "As a patient, I want to receive appointment reminders so I do not miss my visits", "103",
     "- Email reminder sent 24 hours before appointment\n- Reminder includes date, time, doctor name, location\n- Patient can confirm or cancel from the reminder"],

    ["Task", "316", "Create notification service with email support",
     "Implement a notification service that can send templated emails via SMTP", "206", ""],

    ["Task", "317", "Build appointment reminder scheduler",
     "Create a background job that finds appointments 24h away and triggers reminders", "206", ""],

    ["Task", "318", "Create email templates for reminders",
     "Design HTML email templates for appointment reminders with confirm/cancel links", "206", ""],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    row_type = row_data[0]
    if row_type == "Feature":
        fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    elif row_type == "User Story":
        fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    else:
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_idx in range(1, 7):
        ws.cell(row=row_idx, column=col_idx).fill = fill

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 70
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 60

wb.save("sample_backlog.xlsx")
print("Created docs/sample_backlog.xlsx successfully!")
print(f"  - 3 Features")
print(f"  - 6 User Stories")
print(f"  - 18 Tasks")
